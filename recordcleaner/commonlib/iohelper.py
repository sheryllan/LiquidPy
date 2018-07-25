import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter


class XlsxWriter(object):
    @staticmethod
    def load_xlsx(filepath):
        return openpyxl.load_workbook(filepath) if os.path.isfile(filepath) else openpyxl.Workbook()

    @staticmethod
    def create_xlwriter(path, override=True):
        wrt = pd.ExcelWriter(path, engine='openpyxl')

        def __config_xlwriter(wrt, wb):
            wrt.book = wb
            wrt.sheets = dict((ws.title, ws) for ws in wb.worksheets)
            return wrt

        return wrt if override else __config_xlwriter(wrt, XlsxWriter.load_xlsx(path))

    @staticmethod
    def to_xlsheet(data, wrt, sheet, columns=None):
        df = pd.DataFrame(data)
        if columns is not None:
            df = df[columns]
        df.to_excel(wrt, sheet, index=False)

    @staticmethod
    def auto_size_cols(ws):
        i = 0
        ws_cols = list(ws.columns)
        while i < len(ws_cols):
            max_len = max([len(str(row.value)) for row in ws_cols[i]])
            ws.column_dimensions[get_column_letter(i + 1)].width = max_len + 2
            i += 1

    @staticmethod
    def save_sheets(path, sheet2data, columns=None, override=True, auto_size=True):
        outdir = os.path.dirname(path)
        os.makedirs(outdir, exist_ok=True)

        wrt = XlsxWriter.create_xlwriter(path, override)
        for sheet, data in sheet2data.items():
            XlsxWriter.to_xlsheet(data, wrt, sheet, columns)
            if auto_size:
                XlsxWriter.auto_size_cols(wrt.sheets[sheet])
        wrt.save()
        return path


class LogWriter(object):
    def __init__(self, loglevel):
        self.loglevel = loglevel

    def write(self, message):
        message = message.strip()
        if message != '\n':
            self.loglevel(message)

    def flush(self):
        pass

