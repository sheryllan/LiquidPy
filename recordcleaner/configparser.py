import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import tempfile
from utils import *


class XmlParser(object):
    @staticmethod
    # returns a list of dictionaries
    def fltr_attrs(tags, attrs=None):
        if attrs is None:
            return tags
        return [{k: v for k, v in list(tag.attrs.items()) if k in attrs} for tag in tags]

    @staticmethod
    def parse(path, tags=None):
        with open(path, 'r') as input:
            raw_xml = input.read()
        soup = BeautifulSoup(raw_xml, 'html.parser')
        return soup if tags is None else soup.find_all(tags)


class XlsxWriter(object):

    @staticmethod
    def load_xlsx(filepath):
        return openpyxl.load_workbook(filepath) if os.path.isfile(filepath) else openpyxl.Workbook()

    @staticmethod
    def create_xlwriter(path, override=True):
        wrt = pd.ExcelWriter(path, engine='openpyxl')

        def __config_xlwriter(wrt, wb):
            wrt.book = wb
            wrt.sheets = dict((ws.title, ws) for ws in wb.worksheets)
            return wrt

        return wrt if override else __config_xlwriter(wrt, XlsxWriter.load_xlsx(path))

    @staticmethod
    def to_xlsheet(data, wrt, sheet, columns=None):
        df = pd.DataFrame(data) if columns is None else pd.DataFrame(data, columns=columns)
        df.to_excel(wrt, sheet, index=False)

    @staticmethod
    def auto_size_cols(ws):
        i = 0
        while i < ws.max_column:
            max_len = max([len(str(row.value)) for row in list(ws.columns)[i]])
            ws.column_dimensions[get_column_letter(i + 1)].width = max_len + 2
            i += 1

    @staticmethod
    def save_sheets(path, sheet2data, columns=None, override=True, auto_size=True):
        wrt = XlsxWriter.create_xlwriter(path, override)
        for sheet, data in list(sheet2data.items()):
            XlsxWriter.to_xlsheet(data, wrt, sheet, columns)
            if auto_size:
                XlsxWriter.auto_size_cols(wrt.sheets[sheet])
        wrt.save()
        return path


BASE_PATH = '/home/slan/Documents/config_files/'
EXCHANGES = ['asx', 'cme', 'eurex', 'hkfe', 'ice', 'ose', 'sgx']
source_files = {e: e + '.xml' for e in EXCHANGES}

PROPERTIES = {'asx': ['type', 'commodity_code', 'reactor_name'],
              'cme': ['type', 'commodity_name', 'reactor_name'],
              'eurex': ['type', 'commodity', 'reactor_name'],
              'hkfe': ['type', 'market', 'commodity', 'name', 'reactor_name'],
              'ice': ['type', 'market', 'commodity', 'name'],
              'ose': ['type', 'symbol', 'reactor_name', 'description'],
              'sgx': ['typecode', 'symbo', 'reactor_name']
              }
TYPE_COLNUM = 0
# PARAM_KEYS = ['tags', 'fields', 'sheet']
TAGS = ['product']

REPO_URL = 'http://stash.liquid-capital.liquidcap.com/projects/PPT/repos/reactor/browse/files/TNG/products'

INSTRUMENT_TYPES = {'F': 'Futures',
                    'O': 'Options',
                    'S': 'Strategies',
                    'E': 'Equities'}


def get_default_destfiles():
    return {e: BASE_PATH + e + '.xlsx' for e in EXCHANGES}


def get_src_file(exch):
    return BASE_PATH + source_files[exch]


def get_raw_file(filename):
    full_url = '/'.join([REPO_URL, filename])
    f_temp = tempfile.NamedTemporaryFile()
    download(full_url, f_temp)


def parse_config(exch, outpath=None, sheetname=None):
    src_path = get_src_file(exch)
    columns = PROPERTIES[exch]
    data_parsed = XmlParser.fltr_attrs(XmlParser.parse(src_path, TAGS), columns)
    df_parsed = pd.DataFrame(data_parsed, columns=columns)
    type_col = columns[TYPE_COLNUM]
    df_parsed[type_col] = df_parsed[type_col].apply(lambda x: INSTRUMENT_TYPES[x])
    if outpath is not None:
        sheetname = 'configs' if sheetname is None else sheetname
        return XlsxWriter.save_sheets(outpath, {sheetname: df_parsed}, columns, True)
    return df_parsed


def run_all_configs_parse(dest_files=None, sheetname=None):
    results = dict()
    for exch in EXCHANGES:
        outpath = dest_files[exch] if dest_files is not None else None
        df_parsed = parse_config(exch, outpath, sheetname)
        results.update({exch: df_parsed})
    return results


# run_all_configs_parse(get_default_destfiles())

# get_raw_file('asx.xml')
