from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import os


class XmlInputParams(object):
    def __init__(self, path, tags, attrs):
        self.path = path
        self.tags = tags
        self.attrs = attrs


class XlsxOutputParams(object):
    def __init__(self, path, sheet2data):
        self.path = path
        self.sheet2data = sheet2data


class XmlParser(object):
    @staticmethod
    # returns a list of dictionaries
    def fltr_attrs(tags, attrs=None):
        if attrs is None:
            return tags
        return [{k: v for k, v in tag.attrs.items() if k in attrs} for tag in tags]

    @staticmethod
    def parse(path, tags=None):
        with open(path, 'r') as input:
            raw_xml = input.read()
        soup = BeautifulSoup(raw_xml, 'html.parser')
        return soup if tags is None else soup.find_all(tags)


class XlsxWriter(object):

    @staticmethod
    def load_xlsx(filepath):
        # if not os.path.isfile(filepath):
        #     wb = openpyxl.Workbook()
        #     wb.save(filepath)
        # else:
        #     wb = openpyxl.load_workbook(filepath)
        # return wb
        return openpyxl.load_workbook(filepath) if os.path.isfile(filepath) else openpyxl.Workbook()

    @staticmethod
    def __config_xlwriter(wrt, wb):
        wrt.book = wb
        wrt.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        return wrt

    @staticmethod
    def dict_to_xlsheet(data, wrt, sheet, columns=None, override=True):
        df = pd.DataFrame(data) if columns is None else pd.DataFrame(data, columns)
        if not override:
            XlsxWriter.__config_xlwriter()
        df.to_excel(wrt, sheet, index=False)

    @staticmethod
    def parse_save_sheets(inParams, outParam, override=True):
        wrt = pd.ExcelWriter(outParam.path, engine='openpyxl')
        for in_param in zip(inParams, outParam.sheets):
            data_parsed = XmlParser.fltr_attrs(XmlParser.parse(in_param.path, in_param.tags), in_param.attrs)
            XlsxWriter.dict_to_xlsheet(data_parsed, wrt, )
            wrt = pd.ExcelWriter(outputs.path, engine='openpyxl')
        wrt.save()


base_path = '/home/slan/Documents/config_files/'
exchanges = ['asx', 'bloomberg', 'cme', 'eurex', 'hkfe', 'ice', 'ose', 'sgx']
source_files = {e: e + '.xml' for e in exchanges}
dest_files = {e: e + '.xlsx' for e in exchanges}

properties = {'asx': ['type', 'commodity_code', 'reactor_name'],
              'bloomberg': ['reactor_symbol', 'exchange_symbol', 'ISIN', 'market_code'],
              'cme': ['type', 'commodity_code', 'reactor_name'],
              'eurex': ['type', 'commodity', 'reactor_name'],
              'hkfe': ['type', 'market', 'commodity', 'name', 'reactor_name'],
              'ice': ['type', 'market', 'commodity', 'name'],
              'ose': ['type', 'symbol', 'reactor_name'],
              'sgx': ['symbol', 'reactor_name', 'typecode']
              }
param_keys = ['tags', 'fields', 'sheet']
tags = 'product'
sheet = 'Config'


# for exch in exchanges:
#     src = base_path + source_files[exch]
#     dest = base_path + dest_files[exch]
#     params = {param_keys[0]: tags, param_keys[1]: properties[exch], param_keys[2]: sheet}
#     XlsxParser.parse_to_save(src, dest, params)
XmlParser.fltr_attrs()
